from openpyxl import Workbook
import openpyxl
import json
import subprocess

workbook = openpyxl.load_workbook('FBA.xlsx')
worksheet = workbook.get_sheet_by_name("reactions")

# Get iter rows
iter_rows = worksheet.iter_rows()

# Add to row list
row_list = []
for row in iter_rows:
    row_list.append(row)

# Get header list
header_list = []
for header_row in row_list[0]:
    # print(header_row.value)
    header_list.append(header_row.value)

row_list.pop(0)

# Prepare data list
data_list = []

for row in row_list:

    #filter only reaction in cytosol compartment
    if row[2].value.find('[c]') != -1:


        #set arrow style use data in Reversible column , Lower bound column , and Upper bound column
        #row[7] == 1 means reversible column = 1
        #so this reaction can revers and has two arrow at an edge
        source_arrow = ''
        target_arrow = ''
        if row[7].value == 1:
            source_arrow = 'triangle'
            target_arrow = 'triangle'
        #else means row[7] == 0 and reversible column = 0
        #so this reaction can't revers
        #row[8] == 0 means an edge has only arrow at target side
        #lower bound = 0 and upper bound = 1000
        elif row[8].value == 0:
            source_arrow = 'none'
            target_arrow = 'triangle'
        #else row[8] != 0 mean lower bound = 1000 and upper bound = 0
        else :
            source_arrow = 'triangle'
            target_arrow = 'none'

        #get long reaction string store in reaction variable
        reaction = row[2].value
        temp = ''
        
        #split reaction into two group, substance and product
        if reaction.find('->') != -1:
            temp = [reaction[0:reaction.find('->')], reaction[reaction.find('->')+3:]]
        elif reaction.find('<=>') != -1:
            temp = [reaction[0:reaction.find('<=>')], reaction[reaction.find('<=>')+4:]]
        elif reaction.find('<-') != -1:
            temp = [reaction[0:reaction.find('<-')], reaction[reaction.find('<-')+3:]]
        
        #plit substance and product groups into array
        tempS = temp[0].split( )
        tempT = temp[1].split( )

        #set data for each substance side
        for s in tempS:
            if s != '+' and s.find('[c]') != -1:
                cell_dict = {}
                cell_dict['reaction'] = reaction
                #set substance attribute
                cell_dict['substance'] = s
                #Abbreviation column
                cell_dict['abbr'] = row[0].value
                #Name column
                cell_dict['name'] = row[1].value
                #Subsystem column
                cell_dict['pathway'] = row[6].value
                #set arrow style at source side
                cell_dict['source_arrow'] = source_arrow
                #set arrow style at target side
                cell_dict['target_arrow'] = target_arrow
                #Confidence Score column
                cell_dict['flux'] = row[11].value
                data_list.append(cell_dict)

        #set data for each product side
        for t in tempT:
            if t != '+' and t.find('[c]') != -1:
                cell_dict = {}
                cell_dict['reaction'] = reaction
                #set product attribute
                cell_dict['product'] = t
                #Abbreviation column
                cell_dict['abbr'] = row[0].value
                #Name column
                cell_dict['name'] = row[1].value
                #Subsystem column
                cell_dict['pathway'] = row[6].value
                #set arrow style at source side
                cell_dict['source_arrow'] = source_arrow
                #set arrow style at target side
                cell_dict['target_arrow'] = target_arrow
                #Confidence Score column
                cell_dict['flux'] = row[11].value
                data_list.append(cell_dict)

# Convert to JSON
data_json = json.dumps(data_list)

# Write file
with open('result_cytosol.js', 'w') as f:
    f.write('var cytosol = ')
    f.write(data_json)

#subprocess.check_output('start games.html', shell=True)
